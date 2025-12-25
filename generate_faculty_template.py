#!/usr/bin/env python3
"""
Script to generate Excel template for faculty and staff data collection.
This creates a comprehensive template with all fields needed for the faculty data structure.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_faculty_template():
    """Create Excel template for faculty and staff data collection."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Faculty & Staff Data"
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    instruction_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    instruction_font = Font(size=10, italic=True)
    required_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    optional_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Instructions row
    ws.merge_cells('A1:Z1')
    instruction_cell = ws['A1']
    instruction_cell.value = "INSTRUCTIONS: Fill in the data for each faculty/staff member. Required fields are marked with *. For multiple entries (courses, education, etc.), separate with semicolons (;). Example: 'Course 1; Course 2; Course 3'"
    instruction_cell.fill = instruction_fill
    instruction_cell.font = instruction_font
    instruction_cell.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[1].height = 40
    
    # Column headers
    headers = [
        ("ID*", "Unique identifier (e.g., faculty-001)"),
        ("Slug*", "URL-friendly name (e.g., dr-rajesh-sharma)"),
        ("First Name*", "First name"),
        ("Middle Name", "Middle name (optional)"),
        ("Last Name*", "Last name"),
        ("Title", "Dr., Prof., Mr., Ms., etc."),
        ("Designation*", "Professor, Associate Professor, Assistant Professor, Lecturer, etc."),
        ("Department", "Computer Engineering, Biomedical Engineering, etc."),
        ("Faculty Type*", "full-time, part-time, or visiting"),
        ("Category*", "teaching, board-member, administrative, support, or non-teaching"),
        ("Image Path", "Path to image file (e.g., /pp/1.png)"),
        ("Bio", "Biography/description"),
        ("Specialization", "Areas of expertise"),
        ("Courses - Names", "Course names separated by semicolons"),
        ("Courses - Codes", "Course codes separated by semicolons (match order with names)"),
        ("Courses - Programs", "Program names separated by semicolons (match order)"),
        ("Courses - Semesters", "Semester info separated by semicolons (match order)"),
        ("Education - Degrees", "Degrees separated by semicolons (e.g., Ph.D.; M.S.; B.E.)"),
        ("Education - Fields", "Fields of study separated by semicolons"),
        ("Education - Institutions", "Institution names separated by semicolons"),
        ("Education - Years", "Years separated by semicolons"),
        ("Education - Countries", "Countries separated by semicolons"),
        ("Experience - Positions", "Position titles separated by semicolons"),
        ("Experience - Organizations", "Organization names separated by semicolons"),
        ("Experience - Durations", "Duration strings separated by semicolons (e.g., 2015 - Present)"),
        ("Experience - Descriptions", "Descriptions separated by semicolons (optional)"),
        ("Publications - Titles", "Publication titles separated by semicolons"),
        ("Publications - Types", "Types separated by semicolons (journal, conference, book, patent, other)"),
        ("Publications - Venues", "Venue names separated by semicolons"),
        ("Publications - Years", "Years separated by semicolons"),
        ("Publications - Authors", "Author lists separated by semicolons (comma-separated within each)"),
        ("Publications - Links", "URLs separated by semicolons"),
        ("Research Interests - Areas", "Research areas separated by semicolons"),
        ("Research Interests - Descriptions", "Descriptions separated by semicolons (optional)"),
        ("Awards - Titles", "Award titles separated by semicolons"),
        ("Awards - Organizations", "Awarding organizations separated by semicolons"),
        ("Awards - Years", "Years separated by semicolons"),
        ("Awards - Descriptions", "Descriptions separated by semicolons (optional)"),
        ("Contact - Email", "Email address"),
        ("Contact - Phone", "Phone number"),
        ("Contact - Office", "Office location"),
        ("Contact - Website", "Personal/professional website URL"),
        ("Contact - LinkedIn", "LinkedIn profile URL"),
        ("Contact - Google Scholar", "Google Scholar profile URL"),
        ("Contact - ORCID", "ORCID ID"),
        ("Joining Date", "Date in YYYY-MM-DD format"),
        ("Reports To", "ID of supervisor/manager (optional)"),
        ("Manages", "IDs of people they manage, separated by semicolons"),
        ("Board Position", "e.g., Chairman, Member, Secretary"),
        ("Leadership Role", "e.g., Dean, Head of Department, Director"),
        ("Is Active*", "true or false"),
        ("Order", "Display order number (optional)")
    ]
    
    # Write headers
    for col_idx, (header, description) in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.fill = header_fill if "*" in header else (required_fill if "Required" in description else optional_fill)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    
    # Add example row
    example_data = [
        "faculty-001",
        "dr-rajesh-sharma",
        "Rajesh",
        "",
        "Sharma",
        "Dr.",
        "Professor",
        "Computer Engineering",
        "full-time",
        "board-member",
        "/pp/1.png",
        "Dr. Rajesh Sharma is a distinguished professor with over 20 years of experience...",
        "Artificial Intelligence, Machine Learning, Computer Vision",
        "Machine Learning; Neural Networks and Deep Learning; Computer Vision",
        "CS501; CS502; CS601",
        "B. Tech in AI; B. Tech in AI; B. Tech in AI",
        "Semester V; Semester V; Semester VI",
        "Ph.D.; M.S.; B.E.",
        "Computer Science; Computer Science; Computer Engineering",
        "Stanford University; MIT; Tribhuvan University",
        "2005; 2001; 1999",
        "USA; USA; Nepal",
        "Professor; Associate Professor; Research Scientist",
        "NIET; Kathmandu University; Google Research",
        "2015 - Present; 2010 - 2015; 2005 - 2010",
        "Leading research in AI and teaching graduate courses; ; ",
        "Deep Learning for Medical Image Analysis; Neural Architecture Search for Edge Devices",
        "journal; conference",
        "Nature Machine Intelligence; ICML 2022",
        "2023; 2022",
        "R. Sharma, A. Kumar; ",
        "https://example.com/pub1; https://example.com/pub2",
        "Deep Learning; Computer Vision; Medical AI",
        "Neural network architectures and optimization; Image recognition and video analysis; AI applications in healthcare",
        "Best Paper Award; Outstanding Researcher",
        "ICML; Nepal Academy of Science",
        "2022; 2021",
        "; ",
        "info@niet.edu.np",
        "+977-1-XXXXXXX",
        "Block A, Room 301",
        "",
        "linkedin.com/in/rajeshsharma",
        "scholar.google.com/citations?user=rajeshsharma",
        "",
        "2015-01-15",
        "",
        "",
        "Chairman",
        "Head of Department - Computer Engineering",
        "true",
        "1"
    ]
    
    for col_idx, value in enumerate(example_data, start=1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)
        if col_idx <= len(headers):
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Add example label
    example_label = ws.cell(row=3, column=1)
    ws.insert_rows(3)
    example_label = ws.cell(row=3, column=1)
    example_label.value = "EXAMPLE ROW (Delete this row and fill in your data below)"
    example_label.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    example_label.font = Font(bold=True, size=10)
    example_label.alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells(f'A3:{get_column_letter(len(headers))}3')
    
    # Add data validation sheet
    validation_ws = wb.create_sheet("Data Validation")
    validation_ws['A1'] = "Faculty Type Options"
    validation_ws['A2'] = "full-time"
    validation_ws['A3'] = "part-time"
    validation_ws['A4'] = "visiting"
    
    validation_ws['B1'] = "Category Options"
    validation_ws['B2'] = "teaching"
    validation_ws['B3'] = "board-member"
    validation_ws['B4'] = "administrative"
    validation_ws['B5'] = "support"
    validation_ws['B6'] = "non-teaching"
    
    validation_ws['C1'] = "Publication Type Options"
    validation_ws['C2'] = "journal"
    validation_ws['C3'] = "conference"
    validation_ws['C4'] = "book"
    validation_ws['C5'] = "patent"
    validation_ws['C6'] = "other"
    
    validation_ws['D1'] = "Is Active Options"
    validation_ws['D2'] = "true"
    validation_ws['D3'] = "false"
    
    # Format validation sheet
    for col in ['A', 'B', 'C', 'D']:
        cell = validation_ws[f'{col}1']
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        validation_ws.column_dimensions[col].width = 25
    
    # Add guide sheet
    guide_ws = wb.create_sheet("Guide", 0)  # Insert at beginning
    guide_ws['A1'] = "FACULTY & STAFF DATA COLLECTION GUIDE"
    guide_ws['A1'].font = Font(bold=True, size=16)
    guide_ws.merge_cells('A1:D1')
    
    guide_content = [
        "",
        "GENERAL INSTRUCTIONS:",
        "1. Each row represents one faculty or staff member",
        "2. Required fields are marked with *",
        "3. For fields that accept multiple values, separate them with semicolons (;)",
        "4. Keep the order consistent when entering related data (e.g., courses and their codes)",
        "",
        "FIELD-SPECIFIC GUIDELINES:",
        "",
        "ID: Create unique IDs like 'faculty-001', 'staff-001', etc.",
        "Slug: URL-friendly version of name (lowercase, hyphens instead of spaces)",
        "Example: 'Dr. Rajesh Sharma' → 'dr-rajesh-sharma'",
        "",
        "Courses: Enter multiple courses separated by semicolons",
        "  - Names, Codes, Programs, and Semesters must be in the same order",
        "  - Example: 'Machine Learning; Deep Learning' and 'CS501; CS502'",
        "",
        "Education: Enter multiple degrees separated by semicolons",
        "  - All education fields must be in the same order",
        "  - Example: 'Ph.D.; M.S.; B.E.' and 'Computer Science; Computer Science; Engineering'",
        "",
        "Experience: Enter multiple positions separated by semicolons",
        "  - Positions, Organizations, Durations must be in the same order",
        "",
        "Publications: Enter multiple publications separated by semicolons",
        "  - All publication fields must be in the same order",
        "  - For authors, use comma-separated names within each publication",
        "",
        "Research Interests: Enter multiple areas separated by semicolons",
        "",
        "Awards: Enter multiple awards separated by semicolons",
        "",
        "Contact Information: Single values only (one email, one phone, etc.)",
        "",
        "Faculty Type: Must be one of: full-time, part-time, visiting",
        "Category: Must be one of: teaching, board-member, administrative, support, non-teaching",
        "Is Active: Must be 'true' or 'false'",
        "",
        "TIPS:",
        "- Start with a few entries to test the format",
        "- Keep backup copies of your data",
        "- Review the example row for formatting reference",
        "- Contact IT support if you need help with data entry"
    ]
    
    for idx, line in enumerate(guide_content, start=2):
        cell = guide_ws[f'A{idx}']
        cell.value = line
        if line and line.isupper() and not line.startswith("Example"):
            cell.font = Font(bold=True, size=11)
        guide_ws.row_dimensions[idx].height = 20
    
    guide_ws.column_dimensions['A'].width = 80
    guide_ws.column_dimensions['B'].width = 20
    guide_ws.column_dimensions['C'].width = 20
    guide_ws.column_dimensions['D'].width = 20
    
    # Freeze panes on main sheet
    ws.freeze_panes = 'A4'
    
    # Save the file
    filename = f"Faculty_Staff_Data_Template_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"✓ Excel template created: {filename}")
    print(f"  - Main data sheet: 'Faculty & Staff Data'")
    print(f"  - Guide sheet: 'Guide'")
    print(f"  - Validation sheet: 'Data Validation'")
    print(f"\nThe template is ready for data collection!")
    
    return filename

if __name__ == "__main__":
    create_faculty_template()


